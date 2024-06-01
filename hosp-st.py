import base64
import io
import streamlit as st
import pandas as pd
from openpyxl import load_workbook
import pandas as pd
import warnings
warnings.filterwarnings('ignore')
import numpy as np
import cleaner

@st.cache_resource()
def load_data(file, restrict):
    wb = load_workbook(file)
    sheet_names = wb.sheetnames
    dfdict = cleaner.excel_to_dataframes(uploaded_file=file, sheetnames=sheet_names, restrict=restrict)
    cleaned_dfdict = cleaner.validate_all(dfdict=dfdict)
    merged = cleaner.newindex(dfdict=cleaned_dfdict)
    validate_columns = merged.columns[merged.columns.str.endswith('-VALIDATE')]
    print(validate_columns)
    # Flip boolean values in the selected columns
    merged[validate_columns] = ~merged[validate_columns]    
    return merged


def main():
    st.title("Excel File Upload App")
    st.write("This app allows you to upload an Excel file and view its contents.")

    # File upload
    uploaded_file = st.file_uploader("Upload Excel file", type=["xlsx"])
    restrict0 = st.text_input("Hospital type restriction")
    restrict1 = st.text_input("State restriction")
    button = st.button("Submit")
    if uploaded_file is not None and button:
        # Load data using cache
        merged = load_data(uploaded_file, [restrict0,restrict1])
        json_data = merged.to_json(orient='records')

        def download_json(data, filename):
            b64 = base64.b64encode(data.encode()).decode()
            href = f'<a href="data:application/json;base64,{b64}" download="{filename}">Download JSON file</a>'
            return href
        
        for index,row in merged.iterrows():
            if type(row["RATE"]) == str:
                st.write(row["RATE"])
        st.write(merged.iloc[120].to_json())
        st.write(merged.iloc[120])
        st.markdown(download_json(json_data, 'data.json'), unsafe_allow_html=True)

        

if __name__ == "__main__":
    main()
